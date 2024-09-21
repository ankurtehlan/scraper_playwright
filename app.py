import os
import requests
import pandas as pd
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment
from playwright.sync_api import sync_playwright

# Step 1: Define constants
url = "https://www.marutisuzuki.com/genuine-parts/alto-k10-from-aug-2022/2022-till-present/lxi-mt"
PAGE_LIMIT = 34  # Number of pages to scrape

# Step 2: Create lists to store the scraped data
part_numbers = []
part_names = []
mrps = []
image_urls = []

# Step 3: Scrape data using Playwright
def scrape_pages():
    global current_page
    
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)  # Launch Chromium browser
        page = browser.new_page()
        page.goto(url)

        for current_page in range(1, PAGE_LIMIT + 1):
            print(f"Scraping page {current_page}")
            
            # Wait for the product cards to load
            page.wait_for_selector('.sliderBox')
            
            # Get page content and parse with BeautifulSoup
            page_content = page.content()
            soup = BeautifulSoup(page_content, 'html.parser')
            product_cards = soup.find_all('div', class_='sliderBox')
            
            # Extract data from each product card
            for card in product_cards:
                part_number = card.find('p').strong.text 
                part_name = card.find('h3').text.strip()
                mrp = card.find('div', class_='price').text.strip()
                image_url = card.find('img')['src']
                
                part_numbers.append(part_number)
                part_names.append(part_name)
                mrps.append(mrp)
                image_urls.append(image_url)
            
            # Click the 'Next' button if it's available and wait for the next page
            if current_page < PAGE_LIMIT:
                next_button = page.query_selector('.next')
                if next_button:
                    next_button.click()
                    page.wait_for_timeout(3000)  # Wait for the next page to load
                else:
                    print(f"'Next' button not found on page {current_page}")
                    break
        
        browser.close()

# Step 4: Run the scraper function
scrape_pages()

# Step 5: Download the product images to a local directory
if not os.path.exists('images'):
    os.makedirs('images')

image_paths = []
for idx, img_url in enumerate(image_urls):
    img_data = requests.get(img_url).content
    img_filename = f"images/part_image_{idx}.jpg"
    
    with open(img_filename, 'wb') as img_file:
        img_file.write(img_data)
    
    image_paths.append(img_filename)

# Step 6: Create a DataFrame with the scraped data
data = {
    'Part Number': part_numbers,
    'Part Name': part_names,
    'MRP': mrps,
    'Image': image_paths
}
df = pd.DataFrame(data)

# Step 7: Save the DataFrame to an Excel file with embedded images
wb = Workbook()
ws = wb.active
headers = ['Part Number', 'Part Name', 'MRP', 'Image']
ws.append(headers)

# Set column widths and row heights for images
for col in ['A', 'B', 'C', 'D']:
    ws.column_dimensions[col].width = 30

# Write data and insert images into the cells
for i, row in df.iterrows():
    ws.cell(row=i+2, column=1, value=row['Part Number'])
    ws.cell(row=i+2, column=2, value=row['Part Name'])
    ws.cell(row=i+2, column=3, value=row['MRP'])
    
    img = ExcelImage(row['Image'])
    img.height = 100
    img.width = 100
    
    ws.row_dimensions[i+2].height = 100
    ws.add_image(img, f"D{i+2}")
    
    # Center align the image cell
    ws.cell(row=i+2, column=4).alignment = Alignment(horizontal='center', vertical='center')

# Step 8: Save the Excel file with images embedded
wb.save('scraped_parts_with_images_playwright.xlsx')
print("Excel file saved with images!")
